import os
import openpyxl
import random

def generate_labels(root_dir):
    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()

    # 遍历根目录下的所有文件夹
    for sampling_rate_folder in os.listdir(root_dir):
        sampling_rate_path = os.path.join(root_dir, sampling_rate_folder)
        if os.path.isdir(sampling_rate_path):
            # 创建一个新的工作表
            worksheet = workbook.create_sheet(title=sampling_rate_folder)
            worksheet.append(["ID", "Path", "Label", "Label_ID"])

            class_id = 0
            class_map = {}
            file_id = 1
            all_files = []

            for brand_folder in os.listdir(sampling_rate_path):
                brand_path = os.path.join(sampling_rate_path, brand_folder)
                if os.path.isdir(brand_path):
                    for model_folder in os.listdir(brand_path):
                        model_path = os.path.join(brand_path, model_folder)
                        if os.path.isdir(model_path):
                            for phone_folder in os.listdir(model_path):
                                phone_path = os.path.join(model_path, phone_folder)
                                if os.path.isdir(phone_path):
                                    class_label = f"{sampling_rate_folder}_{brand_folder}_{model_folder}_{phone_folder}"
                                    if class_label not in class_map:
                                        class_map[class_label] = class_id
                                        class_id += 1

                                    for txt_file in os.listdir(phone_path):
                                        if txt_file.endswith(".pt"):
                                            txt_file_path = os.path.join(phone_path, txt_file)
                                            all_files.append((file_id, txt_file_path, class_label, class_map[class_label]))
                                            file_id += 1

            # 随机打乱文件列表
            random.shuffle(all_files)

            # 按比例划分训练集和验证集
            train_size = int(0.8 * len(all_files))
            train_files = all_files[:train_size]
            val_files = all_files[train_size:]

            # 添加训练集文件
            for file in train_files:
                worksheet.append([file[0], file[1], f"{file[2]}_train", file[3]])

            # 添加验证集文件
            for file in val_files:
                worksheet.append([file[0], file[1], f"{file[2]}_val", file[3]])

    # 删除默认的工作表
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # 保存工作簿
    workbook.save("labels_half_step.xlsx")

# 使用示例
root_dir = r"G:\dataset\Bluetooth\ADC2IQ_restruct_half_step"
generate_labels(root_dir)
